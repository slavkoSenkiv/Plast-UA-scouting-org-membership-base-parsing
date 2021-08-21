import bs4
import requests
import openpyxl
import time

# <editor-fold desc="start code">


def get_check_write(row, column, selector):
    sheet.cell(row=row, column=1).value = user_id
    sheet.cell(row=row, column=2).value = name
    try:
        value = soup.select(selector)[0].getText()
        if value.endswith('Дата вступу') \
                or value.endswith('Дата Пластової присяги') \
                or value.endswith('Дата Іменування розвідувачем') \
                or value.endswith('Дата Іменування розвідувачем'):
            sheet.cell(row=row, column=column).value = 'нема'
        else:
            if value.startswith('Дата вступу'):
                value = value.replace('Дата вступу', '')
            if value.startswith('Дата Пластової присяги'):
                value = value.replace('Дата Пластової присяги', '')
            if value.startswith('Дата Іменування розвідувачем'):
                value = value.replace('Дата Іменування розвідувачем', '')
            if value.startswith('Дата Іменування розвідувачем'):
                value = value.replace('Дата Іменування розвідувачем', '')
            if value.startswith('Дата Іменування розвідувачкою'):
                value = value.replace('Дата Іменування розвідувачкою', '')
            if value.startswith('Дата початку - '):
                value = value.replace('Дата початку - ', '')
            sheet.cell(row=row, column=column).value = value

    except IndexError:
        sheet.cell(row=row, column=column).value = 'нема'


wb = openpyxl.open('3 проба.xlsx')
sheet = wb.active
print('opening sheet')
url = 'http://3proba.sitegist.net/uk/profile/?userid='
# user_id_list = [2000, 7500, 1000, 7304, 7416]
first_id = 365
last_id = 371  # 7824
cycle_start_time = time.time()

# </editor-fold>


for user_id in range(first_id, last_id + 1):
    print(f'{user_id - (first_id - 1)} / {last_id - first_id}, id = {user_id} ')
    row = user_id - 363
    page = requests.get(url + str(user_id))
    page.raise_for_status()
    soup = bs4.BeautifulSoup(page.text, 'html.parser')

    # <editor-fold desc="scrapping and writing data from profile">
    try:  # if there is no name = empty page
        name = soup.select('div.profileBox > h3')[0].getText()

        get_check_write(row, 3, 'div.panel-body > p:nth-child(1)')  # датаВступу
        get_check_write(row, 4, 'div.panel-body > p:nth-child(2)')  # датаПрисяги
        get_check_write(row, 5, 'div.panel-body > p:nth-child(3)')  # датаІменуванняРозвідувачем
        get_check_write(row, 6, 'div.alert.alert-success')  # часНаПробу
        get_check_write(row, 7, 'div:nth-child(3) > div.panel-body')  # діловедення
        get_check_write(row, 8, 'div:nth-child(4) > div.panel-body')  # табори
        get_check_write(row, 9, 'div:nth-child(5) > div.panel-body')  # пересторогиТаВідзначення
        get_check_write(row, 10, 'div.well > b') # неТретьопробник
        get_check_write(row, 11, 'div.well > div > div.num')  # номерПроби
        get_check_write(row, 12, 'div.well > div > div.date')  # датаПроби
        get_check_write(row, 13, 'div.descriptionCol > h4')  # курінь
        get_check_write(row, 14, '___')  # станиця
        get_check_write(row, 15, '___')  # округа

        vmilosti_string = ''  # вмілості
        vmilosti_element = soup.select('div.uservmilist > div')
        for x in vmilosti_element:
            vmilosti_string += x.getText() + ', '
        sheet.cell(row=row, column=16).value = vmilosti_string.replace('\n', '')

    except IndexError:
        sheet.cell(row=row, column=1).value = user_id
        sheet.cell(row=row, column=2).value = 'пуста сторінка'
    # </editor-fold>

wb.save('3 проба.xlsx')
cycle_time = round(time.time() - cycle_start_time, 2)
print(cycle_time)


