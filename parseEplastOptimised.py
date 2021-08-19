import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException

# <editor-fold desc="start code">
browser = webdriver.Chrome(executable_path=r'/Users/ysenkiv/Code/chromedriver')
wb = openpyxl.open('епласт Львів хлопці oптимізовано.xlsx')
sheet = wb.active


def by_xpath(value_name, column, xpath):
    start_time = time.time()
    print(f'{value_name} ', end='')
    try:
        value = browser.find_element_by_xpath(f'{xpath}').text
    except NoSuchElementException:
        value = f'no {value_name}'
    sheet.cell(row=user_id + 2, column=column).value = value
    print(round(time.time() - start_time, 2), end=' | ')

def by_class(value_name, column, class_name):
    start_time = time.time()
    print(f'{value_name} ', end='')
    try:
        list_of_values = browser.find_elements_by_class_name(f'{class_name}')
        value_string = ''
        for value in list_of_values:
            value_string += ', ' + value.text
    except NoSuchElementException:
        value_string = f'no {value_name}'
    sheet.cell(row=user_id + 2, column=column).value = value_string
    print(round(time.time() - start_time, 2), end=' | ')
# </editor-fold>

# <editor-fold desc="log in into main page">
print('log in into main page')
mainUrl = 'https://plast.sitegist.net/'
browser.get(mainUrl)
loginButton = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'loginLink')))
loginButton.click()
browser.implicitly_wait(10)
loginField = browser.find_element_by_name('f_login_email')
loginField.send_keys('slavko.senkiv@gmail.com')
loginPass = browser.find_element_by_name('f_login_password')
loginPass.send_keys('makeself')
enterButton = browser.find_element_by_xpath('/html/body/div/div[1]/div/div[3]/div[3]/input')
enterButton.click()
# </editor-fold>

# <editor-fold desc="navigating to TA">
print('navigating to TA')
time.sleep(2)
baseButton = browser.find_element_by_class_name('handbook')
baseButton.click()
usersButton = browser.find_element_by_link_text('Користувачі')
usersButton.click()
maleCheck = browser.find_element_by_name('sex-male')
maleCheck.click()
stanucjaField = browser.find_element_by_name('stanucja')
stanucjaField.send_keys('Львів')
time.sleep(1)
lvivButton = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[2]/div[4]/span[1]')
lvivButton.click()
searchButton = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[8]/input')
searchButton.click()
# </editor-fold>

# <editor-fold desc="getting users ids">
print('iterating through TA and scraping them')
users_ids_list = []
usersList = browser.find_elements_by_class_name('user')
for user in usersList:
    user_id = user.get_attribute('joinobjid')
    users_ids_list.append(user_id)
# </editor-fold>

whole_scrapping_cycle_start_time = time.time()

# <editor-fold desc="scrapping TA data">
queue = 1
# for user_id in range(len(users_ids_list)):
for user_id in range(5):

    print(f'\n{queue} / {str(len(users_ids_list))}')
    queue += 1

    # user_id
    sheet.cell(row=user_id + 2, column=1).value = users_ids_list[user_id]
    user_link = 'https://plast.sitegist.net/?pageid=500&userid=' + users_ids_list[user_id]
    browser.get(user_link)

    # <editor-fold desc="name">
    start_time = time.time()
    try:
        name = browser.find_element_by_class_name('profileName').text
    except NoSuchElementException:
        name = 'no name'
    print('name ', end='')
    sheet.cell(row=user_id + 2, column=2).value = name
    print(round(time.time() - start_time, 2), end=' | ')
    # </editor-fold>

    # <editor-fold desc="social">
    start_time = time.time()
    try:
        social = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/a').get_attribute('href')
    except NoSuchElementException:
        social = 'no social'
    print('social ', end='')
    sheet.cell(row=user_id + 2, column=11).value = social
    print(round(time.time() - start_time, 2), end=' | ')
    # </editor-fold>

    by_class('kureni', 4, 'simpleunit')
    by_xpath('stupin', 5, '/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/div[4]/div[2]')
    by_xpath('upu_stupin', 6, '/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/div[15]/div[2]')
    by_class('awards', 10, 'award_item')
    by_class('kvs', 9, 'kv_item')
    by_xpath('education', 12, '/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/div[10]/div[2]')

    # membership
    membershipButton = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[1]/ul/li[2]/a')
    membershipButton.click()

    by_xpath('startdate', 3, '/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[2]/div[1]/div[2]')
    by_xpath('oathdate', 7, '/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[2]/div[2]/div[2]')
    by_xpath('stpl_stupin_date', 8, '/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/table/tbody/tr/td[3]')
    # by_xpath('profession', 13, '/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/div[13]/div[2]')
    by_xpath('2enterDate', 14, '/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[2]')
    # by_xpath('end_of_aprobation', 15, '/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[3]')

# </editor-fold>

wb.save('епласт Львів хлопці oптимізовано.xlsx')
whole_scrapping_cycle_time = round(time.time() - whole_scrapping_cycle_start_time, 2)
print('\n', whole_scrapping_cycle_time)








