import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException

browser = webdriver.Chrome(executable_path=r'/Users/ysenkiv/Code/chromedriver')

# log in into main page
print('log in into main page')
mainUrl = 'https://plast.sitegist.net/'
browser.get(mainUrl)
loginButton = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'loginLink')))
loginButton.click()
browser.implicitly_wait(10)
loginField = browser.find_element_by_name('f_login_email')
loginField.send_keys('slavko.senkiv@gmail.com')
loginPass = browser.find_element_by_name('f_login_password')
loginPass.send_keys('makeself')
enterButton = browser.find_element_by_xpath('/html/body/div/div[1]/div/div[3]/div[3]/input')
enterButton.click()

# navigating to TA
print('navigating to TA')
time.sleep(2)
baseButton = browser.find_element_by_class_name('handbook')
baseButton.click()
usersButton = browser.find_element_by_link_text('Користувачі')
usersButton.click()
maleCheck = browser.find_element_by_name('sex-male')
maleCheck.click()
stanucjaField = browser.find_element_by_name('stanucja')
stanucjaField.send_keys('Львів')
time.sleep(1)
lvivButton = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[2]/div[4]/span[1]')
lvivButton.click()
searchButton = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[8]/input')
searchButton.click()

wb = openpyxl.open('епласт Львів хлопці.xlsx')
sheet = wb.active

# scrapping TA data
print('iterating through TA and scraping them')
users_ids_list = []
usersList = browser.find_elements_by_class_name('user')
for user in usersList:
    user_id = user.get_attribute('joinobjid')
    users_ids_list.append(user_id)

queue = 1
# for id in range(5):
for user_id in range(len(users_ids_list)):

    print(f'{queue} / {str(len(users_ids_list))}')
    queue += 1

    # user_id
    sheet.cell(row=user_id + 2, column=1).value = users_ids_list[user_id]
    user_link = 'https://plast.sitegist.net/?pageid=500&userid=' + users_ids_list[user_id]
    browser.get(user_link)

    # user_name
    try:
        user_name = browser.find_element_by_class_name('profileName').text
    except NoSuchElementException:
        user_name = 'no user_name'
    sheet.cell(row=user_id + 2, column=2).value = user_name

    # user_social
    try:
        user_social = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/a').get_attribute('href')
    except NoSuchElementException:
        user_social = 'no social'
    sheet.cell(row=user_id + 2, column=11).value = user_social

    # user_kureni
    user_kureni = browser.find_elements_by_class_name('simpleunit')
    user_kureni_string = ''
    for kurin in user_kureni:
        user_kureni_string += ', ' + kurin.text
    sheet.cell(row=user_id + 2, column=4).value = user_kureni_string

    try:
        user_stupin = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/div[4]/div[2]').text
    except NoSuchElementException:
        user_stupin = 'no user_stupin'
    sheet.cell(row=user_id + 2, column=5).value = user_stupin

    try:
        user_upu_stupin = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/div[15]/div[2]').text
    except NoSuchElementException:
        user_upu_stupin = 'no user_upu_stupin'
    sheet.cell(row=user_id + 2, column=6).value = user_upu_stupin

    # user_awards
    user_awards = browser.find_elements_by_class_name('award_item')
    user_awards_string = ''
    for award in user_awards:
        user_awards_string += ', ' + award.text
    sheet.cell(row=user_id + 2, column=10).value = user_awards_string

    # kv
    user_kvs = browser.find_elements_by_class_name('kv_item')
    user_kvs_string = ''
    for kv in user_kvs:
        user_kvs_string += ', ' + kv.get_attribute('title')
    sheet.cell(row=user_id + 2, column=9).value = user_kvs_string

    try:
        user_education = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/div[10]/div[2]').text
    except NoSuchElementException:
        user_education = 'no user_education'
    sheet.cell(row=user_id + 2, column=12).value = user_education

    #membership
    membershipButton = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[1]/ul/li[2]/a')
    membershipButton.click()

    # startdate
    try:
        startdate = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[2]/div[1]/div[2]').text
    except NoSuchElementException:
        startdate = 'no startdate'
    sheet.cell(row=user_id + 2, column=3).value = startdate

    # oathdate
    try:
        oathdate = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[2]/div[2]/div[2]').text
    except NoSuchElementException:
        oathdate = 'no oathdate'
    sheet.cell(row=user_id + 2, column=7).value = oathdate

    # stpl_stupin_date
    try:
        stpl_stupin_date = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/table/tbody/tr/td[3]').text
    except NoSuchElementException:
        stpl_stupin_date = 'no stpl_stupin_date'
    sheet.cell(row=user_id + 2, column=8).value = stpl_stupin_date

    # profession
    try:
        profession = browser.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[2]/div[4]/div[1]/div[5]/div[2]/div[13]/div[2]').text
    except NoSuchElementException:
        profession = 'no proffesion'
    sheet.cell(row=user_id + 2, column=13).value = profession

    # additional enterDate
    try:
        enterDate = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[2]').text
    except NoSuchElementException:
        enterDate = 'no additional enterDate'
    sheet.cell(row=user_id + 2, column=13).value = enterDate

    # end of aprobation period
    try:
        end_of_oprobation = browser.find_element_by_xpath('/html/body/div/div[2]/div/div[2]/div[4]/div/div[5]/div[3]').text
    except NoSuchElementException:
        end_of_oprobation = 'no end_of_oprobation'
    sheet.cell(row=user_id + 2, column=13).value = end_of_oprobation

    wb.save('епласт Львів хлопці.xlsx')







